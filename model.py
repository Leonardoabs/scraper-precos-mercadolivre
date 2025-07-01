import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill


def scraper_ml(busca, paginas=1):
    produtos = []
    precos = []

    for pagina in range(1, paginas + 1):
        url = (
            f"https://lista.mercadolivre.com.br/"
            f"{busca}_Desde_{(pagina - 1) * 50 + 1}"
        )

        response = requests.get(url)
        soup = BeautifulSoup(response.text, 'html.parser')

        itens = soup.find_all('li', class_='ui-search-layout__item')

        for item in itens:
            titulo_tag = item.find('a', class_='poly-component__title')
            preco_tag = item.find(
                'span', class_='andes-money-amount__fraction')

            if titulo_tag and preco_tag:
                produtos.append(titulo_tag.text.strip())
                precos.append("R$ " + preco_tag.text.strip())

    df = pd.DataFrame({'Produto': produtos, 'Preço (R$)': precos})

    arquivo = f"produtos_{busca}.xlsx"
    df.to_excel(arquivo, index=False)

    # ESTILIZAÇÃO PLANILHA
    from openpyxl import load_workbook

    wb = load_workbook(arquivo)
    ws = wb.active

    # Estilos
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    align_left = Alignment(horizontal="left", vertical="center")

    # ESTILHO NO CABEÇALHO
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_left

    # ALINHAMENTO CELULAS
    for row in ws.iter_rows(min_row=2,
                            max_row=ws.max_row, min_col=1, max_col=2):

        for cell in row:
            cell.alignment = align_left

    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 15

    wb.save(arquivo)
    return arquivo
